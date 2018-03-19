import openpyxl
import get_host
import json
import extool
import os
import put_host
import json
import ruleNote
import get_rules
import DE
import logging
from pprint import pprint
from Log import DBTool
import os
import sys


#確認檔案存在是否,沒有就創建一個檔案
def ckfile():
    if os.path.isfile("test1.xlsx"):
        workbook = openpyxl.load_workbook("test1.xlsx")
    else:
        workbook = extool.create()
    return workbook

def ckuser():
    filepath = 'user.json'
    if os.path.isfile(filepath):
        user = open("user.json", "r")
        user = json.loads(user.read())
        
        return  user
        
    else:
         user = open("user.json", "rw")
         return user
        


#更新處理
def putloop(workbook, sheet, fields ,obj):
    setdata = {}
    for row in range(2,sheet.max_row + 1):
        for field in fields:
            setdata[field] = extool.gett(sheet, row, fields.index(field) + 1)
            
        get_id = setdata['id']
        put_host.putdata(setdata, get_id, obj)
        setdata = {}
    print("OK!!!!")
        

#取得處理  
def getloop(workbook,sheet ,data, field):
    index = 2
    inda = 0
    base = data.json()['items']
    while inda < len(base) :
        for get_data in field:
            try:
                extool.write(workbook, sheet, index,field.index(get_data)+ 1, base[inda][get_data] )
            except KeyError:
                print("havn\'t", get_data, "in" ,base[index]['id'], "list.")
                continue
        index += 1
        inda  += 1
    print("OK!!!!")
    

#取得資料
def get_datas(obj):
    if obj == 'protocolportobjects':
        field = get_host.set_field_protocol()
    else:
        field = get_host.set_field()
        
    sheet = workbook.get_sheet_by_name(obj)
    for x in field:
        extool.write(workbook, sheet, 1, field.index(x)+1,x)  
    host = get_host.a(obj)
    getloop(workbook,sheet,host,field)
    
#更新資料
def put_datas(obj):
    sheet = workbook.get_sheet_by_name(obj)
    field = put_host.set_field()
    putloop(workbook, sheet, field, obj)

#刪除資料
def del_datas():
    sheet = workbook.get_sheet_by_name('De')
    for row in range(2,sheet.max_row + 1):
        get_id = extool.gett(sheet, row, 2)
        get_type = extool.gett(sheet, row, 1)
        DE.a(get_type, get_id)
        
    

#取得ruleNote
def getruleNote():
    ruleNote.a()
    
#取得rule
def getrules(get_id):
    get_rules.a(get_id)
    
#Object選單
def listobj():
    sheet_na = ['hosts', 'networks', 'protocolportobjects', 'ranges']
    print("選項清單:")
    for na in sheet_na:
        print(sheet_na.index(na) + 1,":",na)

    a = input("選擇選項----->")
    a = int(a) -1
    print(sheet_na[a])        
    return  (sheet_na[a])



    
def meau():
    print("--------------FTD API 初版----------------------------")
    print("請選擇你要的功能\n(1)取得Object\n(2)修正Object(需有excel檔)\n(3)刪除(尚未開放)\n(4)退出")    
    ch1 = input("在此輸入 ----->")


    if ch1 == '1':
        ch2 = listobj()
        get_datas(ch2)
        DBTool()
        meau()
    elif ch1 == '2':
        ch2 = listobj()
        put_datas(ch2)
        meau()
    elif ch1 == '3':
        del_datas()
        meau()
    elif ch1 == '4':
        sys.exit()
    else:
        print("沒有此功能!!!")
        meau()
workbook = ckfile()




meau()









    
    












