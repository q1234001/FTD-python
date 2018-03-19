import openpyxl


def create():
    workbook = openpyxl.Workbook()
    sheet_na = ['hosts', 'networks', 'protocolportobjects', 'ranges', 'De']
    
    for na in sheet_na:
        workbook.create_sheet(str(na))

    workbook.save('test1.xlsx')
    return workbook
    
        
def write(workbook, sheet, x, y, ans):
    sheet.cell(row = x, column = y ).value = ans
    workbook.save('test1.xlsx')


def gett(sheet,x,y):
    return sheet.cell(row = x, column = y ).value
    

    
